'''
Created on 15-Mar-2017

@author: sidtyagi
'''
'''
Created on 14-Mar-2017

@author: sidtyagi
'''


import xlrd,openpyxl,time
from openpyxl.styles import Alignment,Font,colors
import os

def open_target_workbook(workbook_name):
    xfile_original = openpyxl.load_workbook(workbook_name)
    return xfile_original



#def edit_the_workbook(row,col,config,sheet_name,xfile_original):
def edit_the_workbook(row,col,sheet_name,xfile_original):
    print 'writing to row '+str(row)
    print 'writing to col '+str(col)
    #----------------------------#
    #xfile_original = openpyxl.load_workbook('TEST_WITH_BLANKS - Copy.xlsx')  
    
    new_sheet = xfile_original.get_sheet_by_name(sheet_name)
    
    #print dir(new_sheet.cell(row=row, column=col))
    
    
    new_sheet.cell(row=row, column=col).font=Font(color=colors.RED)
    #new_sheet.cell(row=1, column=1).font=Font(color=colors.RED)
    #new_sheet.cell(row=row, column=col).value = config
    
    #xfile_original.save('TEST_WITH_BLANKS - Color.xlsx') 
def save_the_workbook(handler):
    handler.save('TEST_WITH_BLANKS - Color.xlsx')
    



#--------open the target workbook where u will insert colors------------#
target_excel_handler=open_target_workbook('TEST_WITH_BLANKS - Color.xlsx')



#---open the original workbook-----#
xl_copy = xlrd.open_workbook(r'TEST_WITH_BLANKS - Copy.xlsx', on_demand=True)
#print xl_copy.sheet_names()

all_sheets = xl_copy.sheet_names()
print all_sheets
for sheet in all_sheets:  
    print '**********Now doing sheet*************'
    print str(sheet)
    print '***********************'
    xl_sheet_copy=''
    xl_sheet_copy = xl_copy.sheet_by_name(sheet)    
    row_count=xl_sheet_copy.nrows
    col_count=xl_sheet_copy.ncols
    print 'row count is'
    print row_count
    print ' col_count is ' 
    print col_count
    for x in  range(0,row_count):  
        for y in range(5,col_count):             
            cell_val=xl_sheet_copy.cell(x,y).value 
            cell_val=str(cell_val)
            print type(cell_val)
            print cell_val+'TTTTTTTTTTTt'            
            if 'MUM-MPL-PE-RTR-37-241' in cell_val.strip() or 'MUM-MPL-PE-RTR-37-211' in cell_val.strip():
            
            #if '202.123.37.35' in cell_val.strip():
            #    print 'coloring --------------------------------'
            #if cell_val in dict_ip_mapping:
                #row_no_for_new_sheet=x+2
                row_no_for_new_sheet=x+1
                #config=dict_ip_mapping[cell_val]
                #edit_the_workbook(row_no_for_new_sheet,y+1,config,sheet)
                
                #edit_the_workbook(row_no_for_new_sheet,y+1,config,sheet,target_excel_handler)
                #edit_the_workbook(row_no_for_new_sheet,y+1,config,sheet,target_excel_handler)
                edit_the_workbook(row_no_for_new_sheet,y+1,sheet,target_excel_handler)
        
save_the_workbook(target_excel_handler)
