'''
@author: sidtyagi
'''

import openpyxl,xlrd
from openpyxl.styles import Font,colors




def create_workbook(book_name):
    #create a new workbook
    wb = openpyxl.Workbook()
    wb.save(book_name)


def add_sheet_to_workbook(book_name):
    #---add a sheet to an existing workbook--#
    wb = openpyxl.load_workbook(book_name)
    #wb.create_sheet("Sheet2")
    #wb.create_sheet("test_sheet")
    wb.save(book_name)

def get_all_sheets(book_name):
    #xl_copy = xlrd.open_workbook(r'TEST_WITH_BLANKS - Copy.xlsx', on_demand=True)
    #xl_copy = xlrd.open_workbook(book_name, on_demand=True)
    xl_copy = xlrd.open_workbook(book_name)
    #print xl_copy.sheet_names()
    sheets = xl_copy.sheet_names()
    print sheets


def read_from_cells(book_name):
    xl_copy = xlrd.open_workbook(book_name)
    xl_sheet_copy = xl_copy.sheet_by_name('test_sheet')    
    row_count=xl_sheet_copy.nrows
    col_count=xl_sheet_copy.ncols
    print 'the row count %s'%row_count
    print 'the col count %s'%col_count

    #read first cell i.e 0th row, 0th column
    cell_val=xl_sheet_copy.cell(0,0).value
    print cell_val

    #read last cell i.e 2nd row , 1st column
    cell_val=xl_sheet_copy.cell(2,1).value
    print cell_val
    
def write_to_cell(book_name,row,col):
    print 'writing to row '+str(row)
    print 'writing to col '+str(col)
    #----------------------------#
    xfile_original = openpyxl.load_workbook(book_name)  
    
    #new_sheet = xfile_original.get_sheet_by_name(sheet_name)
    sheet = xfile_original.get_sheet_by_name('test_sheet')    
    sheet.cell(row=row, column=col).font=Font(color=colors.RED)
    #new_sheet.cell(row=1, column=1).font=Font(color=colors.RED)
    #new_sheet.cell(row=row, column=col).value = config
    sheet.cell(row=row, column=col).value = 'Pushed this text via script'
    xfile_original.save(book_name) 

#create_workbook('test2.xlsx')
#add_sheet_to_workbook('test2.xlsx')
get_all_sheets('test2.xlsx')
read_from_cells('test2.xlsx')
write_to_cell('test2.xlsx',5,4)
